#!/usr/bin/env python3
"""
AFL Dashboard – Scheduled Data Update
======================================
Runs all scrapers and processing steps to keep the dashboard current
during the 2026 AFL season.

Designed to be executed automatically via cron/launchd every Monday and Friday,
or manually at any time.

Usage:
    python scheduled_update.py              # Full update (all steps)
    python scheduled_update.py --quick      # Quick update (skip slow scrapers)
    python scheduled_update.py --dry-run    # Show what would run without executing

What it does (in order):
  1. Scrape player lists from Footywire (squad rosters, DOBs, heights)
  2. Scrape extended data from Footywire (contracts, draft history)
  3. Scrape AFL ladder from Footywire (season standings)
  4. Scrape Wheelo Ratings (team + player stats) via Selenium
  5. Fetch Traits API data for players
  6. Build season data files (squads, player_stats, traits CSVs)
  7. Regenerate computed ratings (team summaries, ladders)
  8. Auto-update player photos for any new/missing players
  9. Validate file freshness
  10. Validate CSV schemas

Logs are saved to: logs/scheduled_update_YYYY-MM-DD.log
"""

import os
import sys
import time
import logging
import argparse
import traceback
import subprocess
from pathlib import Path
from datetime import datetime

# ============================================================================
# CONFIGURATION
# ============================================================================
BASE_DIR = Path(__file__).parent
LOG_DIR = BASE_DIR / "logs"
VENV_PYTHON = BASE_DIR / ".venv" / "bin" / "python"  # Fallback if conda not available
CONDA_ENV = "afl"

# Determine which Python to use
def get_python():
    """Get the correct Python executable (conda preferred, .venv fallback)."""
    # If we're already in the conda env, just use current python
    if os.environ.get("CONDA_DEFAULT_ENV") == CONDA_ENV:
        return sys.executable
    # If .venv exists, use it
    if VENV_PYTHON.exists():
        return str(VENV_PYTHON)
    # Fallback to current python
    return sys.executable


# Steps to run, in order.
# Each step is (name, script_path, args, description, slow?)
UPDATE_STEPS = [
    (
        "footywire_squads",
        "scrape_footywire.py",
        [],
        "Scrape player squads from Footywire",
        False,
    ),
    (
        "footywire_extended",
        "scrape_footywire_extended.py",
        ["--current-only"],  # Fast mode: only current season drafts, use draft cache
        "Scrape contracts & draft history from Footywire",
        False,
    ),
    (
        "afl_ladders",
        "scrape_afl_ladders.py",
        ["--current-only"],  # Incremental: only scrape current season, merge with history
        "Scrape AFL ladder standings from Footywire (current season)",
        False,
    ),
    (
        "wheelo_ratings",
        "scrape_wheelo_ratings.py",
        ["--all"],
        "Scrape Wheelo team & player ratings (Selenium)",
        True,  # This is slow — uses Selenium
    ),
    (
        "traits_api",
        "run_traits_api.py",
        [],
        "Fetch Traits API data for current-season players",
        True,  # This hits an external API
    ),
    (
        "wheelo_to_raw",
        None,  # Inline step, not a script
        [],
        "Copy Wheelo team stats to raw data folder for ratings pipeline",
        False,
    ),
    (
        "build_season",
        "build_season_data.py",
        [],
        "Build season data files (squads, stats, traits CSVs)",
        False,
    ),
    (
        "wheelo_player_to_raw",
        None,  # Inline step, not a script
        [],
        "Merge Wheelo player season data (matches, ratings) into raw player CSVs",
        False,
    ),
    (
        "regenerate_ratings",
        "regenerate_ratings.py",
        [],
        "Regenerate computed team summaries and ladders",
        False,
    ),
    (
        "refresh_master_ladders",
        None,  # Inline step
        [],
        "Update master workbook Team_Ladders_All with freshly-scraped ladder data",
        False,
    ),
    (
        "update_photos",
        "auto_update_photos.py",
        [],
        "Download photos for any new/missing players",
        True,
    ),
    (
        "validate_freshness",
        "validate_pipeline_freshness.py",
        [],
        "Validate pipeline output files are present and fresh",
        False,
    ),
    (
        "validate_schemas",
        None,  # Inline step
        [],
        "Validate CSV output schemas (columns, row counts, data quality)",
        False,
    ),
    (
        "data_diff_check",
        None,  # Inline step
        [],
        "Compare pipeline outputs against backups and alert on anomalies",
        False,
    ),
]

# Dependency chain: if a prerequisite fails, the dependent step is SKIPPED.
# Steps not listed here have no dependencies and always run.
STEP_DEPENDENCIES = {
    "build_season":          ["footywire_squads"],
    "wheelo_to_raw":         ["wheelo_ratings"],
    "wheelo_player_to_raw":  ["wheelo_ratings", "build_season"],
    "regenerate_ratings":    ["wheelo_to_raw"],
    "refresh_master_ladders":["afl_ladders", "regenerate_ratings"],
    "validate_schemas":      ["build_season"],
    "data_diff_check":       ["build_season"],
}

# Maximum number of scheduled log files to keep
MAX_LOG_FILES = 30


# ============================================================================
# LOGGING
# ============================================================================
def setup_logging():
    """Configure logging to file and console."""
    LOG_DIR.mkdir(exist_ok=True)
    log_file = LOG_DIR / f"scheduled_update_{datetime.now():%Y-%m-%d_%H%M}.log"

    fmt = "%(asctime)s [%(levelname)s] %(message)s"
    logging.basicConfig(
        level=logging.INFO,
        format=fmt,
        handlers=[
            logging.FileHandler(log_file),
            logging.StreamHandler(sys.stdout),
        ],
    )
    logging.info(f"Log file: {log_file}")
    return log_file


def rotate_logs():
    """Remove old scheduled-update log files, keeping the most recent MAX_LOG_FILES."""
    log_files = sorted(
        LOG_DIR.glob("scheduled_update_*.log"),
        key=lambda p: p.stat().st_mtime,
        reverse=True,
    )
    removed = 0
    for old_log in log_files[MAX_LOG_FILES:]:
        try:
            old_log.unlink()
            removed += 1
        except OSError:
            pass
    if removed:
        logging.info(f"Log rotation: removed {removed} old log file(s)")


def send_macos_notification(title, message, sound="Basso"):
    """Send a macOS notification via osascript. Silent no-op on non-Mac.
    DEPRECATED: Use utils.notifications.notify() instead for unified notifications.
    """
    try:
        from utils.notifications import send_macos_notification as _send
        _send(title, message, sound)
    except ImportError:
        try:
            import subprocess as _sp
            _sp.run(
                [
                    "osascript", "-e",
                    f'display notification "{message}" with title "{title}" sound name "{sound}"',
                ],
                timeout=5,
                capture_output=True,
            )
        except Exception:
            pass


def _notify(title, message, is_error=False):
    """Send notification to all configured channels (macOS + email)."""
    try:
        from utils.notifications import notify
        notify(title, message, is_error=is_error)
    except ImportError:
        send_macos_notification(title, message)


# ============================================================================
# INLINE STEPS (no external script needed)
# ============================================================================
def run_inline_step(name, description):
    """Run an inline processing step."""
    logging.info(f"START [{name}] {description}")
    start = time.time()

    try:
        if name == "wheelo_to_raw":
            import pandas as pd
            wheelo_path = BASE_DIR / "Wheelo_Team_Data.xlsx"
            if not wheelo_path.exists():
                logging.warning(f"  SKIP [{name}] – Wheelo_Team_Data.xlsx not found")
                return True, None  # Not a failure, just nothing to copy

            xl = pd.ExcelFile(wheelo_path)
            import re
            # Process Season, L10, and L5 sheets
            for sheet in xl.sheet_names:
                df = xl.parse(sheet)
                # Remove summary rows (like 'Average')
                df = df[df['Team'].notna()]
                df = df[~df['Team'].astype(str).str.contains('Average|Total|nan', case=False, na=False)]
                # Extract year from sheet name (e.g. "Wheelo 2026 Season" -> 2026)
                year_match = re.search(r'(\d{4})', sheet)
                if not year_match:
                    continue
                year = year_match.group(1)
                # Determine suffix based on sheet type
                if "L10" in sheet:
                    suffix = "_L10"
                elif "L5" in sheet:
                    suffix = "_L5"
                elif "Season" in sheet:
                    suffix = ""
                else:
                    continue
                from utils.safe_io import safe_csv_write
                out_path = BASE_DIR / "data" / "raw" / "team" / f"team_stats_{year}{suffix}.csv"
                safe_csv_write(df, out_path)
                logging.info(f"  Copied {sheet} -> {out_path.name} ({len(df)} teams)")

            elapsed = time.time() - start
            logging.info(f"  DONE [{name}] in {elapsed:.1f}s")
            return True, None
        elif name == "refresh_master_ladders":
            import pandas as pd
            from config.constants import MASTER_FILE, LADDERS_FILE
            master_path = BASE_DIR / MASTER_FILE
            ladder_path = BASE_DIR / LADDERS_FILE
            if not master_path.exists():
                logging.warning(f"  SKIP [{name}] – master workbook not found")
                return True, None
            if not ladder_path.exists():
                logging.warning(f"  SKIP [{name}] – ladder file not found")
                return True, None
            ladder_df = pd.read_excel(ladder_path)
            from openpyxl import load_workbook
            wb = load_workbook(master_path)
            if "Team_Ladders_All" in wb.sheetnames:
                del wb["Team_Ladders_All"]
            wb.save(master_path)
            wb.close()
            with pd.ExcelWriter(master_path, engine="openpyxl", mode="a", if_sheet_exists="replace") as writer:
                ladder_df.to_excel(writer, sheet_name="Team_Ladders_All", index=False)
            logging.info(f"  Refreshed Team_Ladders_All: {len(ladder_df)} rows")
            elapsed = time.time() - start
            logging.info(f"  DONE [{name}] in {elapsed:.1f}s")
            return True, None

        elif name == "wheelo_player_to_raw":
            import pandas as pd
            import re as re_mod
            wheelo_path = BASE_DIR / "Wheelo_Player_Data.xlsx"
            if not wheelo_path.exists():
                logging.warning(f"  SKIP [{name}] – Wheelo_Player_Data.xlsx not found")
                return True, None

            # ── Team name normalisation ──────────────────────────────
            # Maps every known variant to the canonical name used in
            # Footywire / app data.  Extend as new variants appear.
            TEAM_NAME_MAP = {
                "Greater Western Sydney": "GWS Giants",
                "GWS":                    "GWS Giants",
                "Brisbane Lions":         "Brisbane",
                "Adelaide Crows":         "Adelaide",
                "Sydney Swans":           "Sydney",
                "Geelong Cats":           "Geelong",
                "West Coast Eagles":      "West Coast",
                "Gold Coast Suns":        "Gold Coast",
                "North Melbourne Kangaroos": "North Melbourne",
                "Carlton Blues":          "Carlton",
                "Collingwood Magpies":    "Collingwood",
                "Essendon Bombers":       "Essendon",
                "Fremantle Dockers":      "Fremantle",
                "Hawthorn Hawks":         "Hawthorn",
                "Melbourne Demons":       "Melbourne",
                "Port Adelaide Power":    "Port Adelaide",
                "Richmond Tigers":        "Richmond",
                "St Kilda Saints":        "St Kilda",
                "Western Bulldogs":       "Western Bulldogs",
            }

            # ── Multi-tier player matching ───────────────────────────
            def _build_player_map(wheelo_df, target_df):
                """Build a mapping of Wheelo player names → target CSV names.

                Uses a 3-tier matching strategy:
                  Tier 1: Exact Player + Team  (catches ~90%)
                  Tier 2: Same surname + same team, only one candidate
                  Tier 3: Same surname + same team, disambiguate by
                           closest Age_Decimal

                Returns dict  {wheelo_name: target_name}  for mismatched
                names only (exact matches don't need remapping).
                Also returns a list of unmatched Wheelo player names.
                """
                import difflib

                w_set = set(zip(wheelo_df["Player"], wheelo_df["Team"]))
                t_set = set(zip(target_df["Player"], target_df["Team"]))

                # Tier 1 – exact match
                exact = w_set & t_set
                unmatched_w = w_set - exact
                unmatched_t = t_set - w_set  # target names not yet claimed

                if not unmatched_w:
                    return {}, []

                # Pre-index target by (surname_lower, team)
                from collections import defaultdict
                t_by_surname = defaultdict(list)
                for tn, tt in unmatched_t:
                    surname = tn.split()[-1].lower()
                    t_by_surname[(surname, tt)].append(tn)

                # Build Age_Decimal lookup for disambiguation
                w_age = {}
                if "Age_Decimal" in wheelo_df.columns:
                    for _, r in wheelo_df.iterrows():
                        w_age[(r["Player"], r["Team"])] = r["Age_Decimal"]
                t_age = {}
                if "Age_Decimal" in target_df.columns:
                    for _, r in target_df.iterrows():
                        t_age[(r["Player"], r["Team"])] = r["Age_Decimal"]
                elif "Age" in target_df.columns:
                    for _, r in target_df.iterrows():
                        try:
                            t_age[(r["Player"], r["Team"])] = float(r["Age"])
                        except (ValueError, TypeError):
                            pass

                nick_map = {}
                still_unmatched = []

                for wn, wt in unmatched_w:
                    surname = wn.split()[-1].lower()
                    candidates = t_by_surname.get((surname, wt), [])

                    if len(candidates) == 1:
                        # Tier 2 – unique surname on same team
                        nick_map[wn] = candidates[0]
                        # Remove from pool so it can't be claimed again
                        t_by_surname[(surname, wt)] = []

                    elif len(candidates) > 1:
                        # Tier 3a – try difflib on full name first
                        best = difflib.get_close_matches(
                            wn, candidates, n=1, cutoff=0.6)
                        if best:
                            nick_map[wn] = best[0]
                            candidates.remove(best[0])
                        else:
                            # Tier 3b – closest Age_Decimal
                            w_a = w_age.get((wn, wt))
                            if w_a is not None:
                                scored = []
                                for cn in candidates:
                                    t_a = t_age.get((cn, wt))
                                    if t_a is not None:
                                        scored.append((abs(w_a - t_a), cn))
                                if scored:
                                    scored.sort()
                                    # Only accept if age difference ≤ 1 year
                                    if scored[0][0] <= 1.0:
                                        nick_map[wn] = scored[0][1]
                                        candidates.remove(scored[0][1])
                                    else:
                                        still_unmatched.append((wn, wt))
                                else:
                                    still_unmatched.append((wn, wt))
                            else:
                                still_unmatched.append((wn, wt))
                    else:
                        # No surname match at all on this team
                        still_unmatched.append((wn, wt))

                return nick_map, still_unmatched

            xl = pd.ExcelFile(wheelo_path)
            # Process each "Wheelo {year} Season" sheet
            for sheet in xl.sheet_names:
                if "Season" not in sheet:
                    continue
                year_match = re_mod.search(r'(\d{4})', sheet)
                if not year_match:
                    continue
                year = year_match.group(1)

                wheelo_df = xl.parse(sheet)
                # Clean up summary rows
                wheelo_df = wheelo_df[wheelo_df['Player'].notna()]
                wheelo_df = wheelo_df[~wheelo_df['Player'].astype(str).str.contains(
                    'Average|Total|nan', case=False, na=False)]
                wheelo_df["Player"] = wheelo_df["Player"].astype(str).str.strip()
                wheelo_df["Team"] = wheelo_df["Team"].astype(str).str.strip()
                # Normalise team names to match Footywire/app conventions
                wheelo_df["Team"] = wheelo_df["Team"].replace(TEAM_NAME_MAP)
                logging.info(f"  Wheelo {sheet}: {len(wheelo_df)} players, "
                             f"{wheelo_df['Team'].nunique()} teams")

                # --- Update player_stats_{year}.csv ---
                stats_path = BASE_DIR / "data" / "raw" / "player" / f"player_stats_{year}.csv"
                if stats_path.exists():
                    stats_df = pd.read_csv(stats_path)
                    stats_df["Player"] = stats_df["Player"].astype(str).str.strip()
                    stats_df["Team"] = stats_df["Team"].astype(str).str.strip()

                    # Build player name mapping
                    nick_map, unmatched = _build_player_map(wheelo_df, stats_df)
                    if nick_map:
                        logging.info(f"  Name mappings ({len(nick_map)}): {nick_map}")
                    if unmatched:
                        logging.warning(f"  UNMATCHED Wheelo players ({len(unmatched)}): "
                                        f"{unmatched}")

                    # Apply nickname mapping to wheelo data before merge
                    wheelo_work = wheelo_df.copy()
                    wheelo_work["Player"] = wheelo_work["Player"].replace(nick_map)

                    # Columns to update from Wheelo
                    update_cols = [c for c in [
                        "Matches", "RatingPoints_Avg", "CoachesVotes_Avg",
                        "CoachesVotes_Total", "TimeOnGround", "Supercoach_Avg",
                        "DreamTeamPoints_Avg",
                    ] if c in wheelo_work.columns]

                    wheelo_merge = wheelo_work[["Player", "Team"] + update_cols].copy()
                    wheelo_merge = wheelo_merge.rename(
                        columns={c: f"{c}_wheelo" for c in update_cols})

                    stats_df = stats_df.merge(
                        wheelo_merge, on=["Player", "Team"], how="left")

                    # Overwrite with Wheelo values where available
                    for col in update_cols:
                        wcol = f"{col}_wheelo"
                        if wcol in stats_df.columns:
                            mask = stats_df[wcol].notna()
                            if col not in stats_df.columns:
                                stats_df[col] = 0
                            stats_df.loc[mask, col] = stats_df.loc[mask, wcol]
                            stats_df.drop(columns=[wcol], inplace=True)

                    from utils.safe_io import safe_csv_write as _scsv
                    _scsv(stats_df, stats_path)
                    matched_count = (stats_df["Matches"] > 0).sum() if "Matches" in stats_df.columns else 0
                    logging.info(f"  Updated player_stats_{year}.csv: "
                                 f"{matched_count} players with matches from Wheelo")

                # --- Update squads_{year}.csv (Matches_Current) ---
                squads_path = BASE_DIR / "data" / "raw" / "player" / f"squads_{year}.csv"
                if squads_path.exists():
                    squads_df = pd.read_csv(squads_path)
                    squads_df["Player"] = squads_df["Player"].astype(str).str.strip()
                    squads_df["Team"] = squads_df["Team"].astype(str).str.strip()

                    if "Matches" in wheelo_df.columns:
                        # Build a fresh nick_map for squads if needed
                        sq_nick_map, sq_unmatched = _build_player_map(
                            wheelo_df, squads_df)
                        wheelo_sq = wheelo_df.copy()
                        wheelo_sq["Player"] = wheelo_sq["Player"].replace(sq_nick_map)
                        matches_merge = wheelo_sq[["Player", "Team", "Matches"]].copy()
                        matches_merge.rename(
                            columns={"Matches": "Matches_wheelo"}, inplace=True)
                        squads_df = squads_df.merge(
                            matches_merge, on=["Player", "Team"], how="left")
                        mask = squads_df["Matches_wheelo"].notna()
                        if "Matches_Current" not in squads_df.columns:
                            squads_df["Matches_Current"] = 0
                        squads_df.loc[mask, "Matches_Current"] = \
                            squads_df.loc[mask, "Matches_wheelo"]
                        squads_df.drop(columns=["Matches_wheelo"], inplace=True)
                        from utils.safe_io import safe_csv_write as _scsv2
                        _scsv2(squads_df, squads_path)
                        matched_count = mask.sum()
                        logging.info(f"  Updated squads_{year}.csv: "
                                     f"{matched_count} players matched from Wheelo")

            elapsed = time.time() - start
            logging.info(f"  DONE [{name}] in {elapsed:.1f}s")
            return True, None

        elif name == "validate_schemas":
            from utils.schema_validator import validate_pipeline_schemas
            from config.constants import CURRENT_SEASON
            errors = validate_pipeline_schemas(CURRENT_SEASON)
            warnings = [e for e in errors if e.is_warning]
            hard_errors = [e for e in errors if not e.is_warning]
            for w in warnings:
                logging.warning(f"  {w}")
            for e in hard_errors:
                logging.error(f"  {e}")
            elapsed = time.time() - start
            if hard_errors:
                msg = f"{len(hard_errors)} schema error(s), {len(warnings)} warning(s)"
                logging.error(f"  FAILED [{name}] {msg} ({elapsed:.1f}s)")
                return False, msg
            logging.info(f"  DONE [{name}] {len(warnings)} warning(s) in {elapsed:.1f}s")
            return True, None

        elif name == "data_diff_check":
            from utils.data_diff import diff_report
            from config.constants import CURRENT_SEASON

            files_to_check = [
                BASE_DIR / "data" / "raw" / "player" / f"squads_{CURRENT_SEASON}.csv",
                BASE_DIR / "data" / "raw" / "player" / f"player_stats_{CURRENT_SEASON}.csv",
                BASE_DIR / "data" / "raw" / "player" / f"footywire_{CURRENT_SEASON}_complete.csv",
                BASE_DIR / "data" / "raw" / "traits" / f"traits_{CURRENT_SEASON}.csv",
                BASE_DIR / "data" / "computed" / f"team_summary_{CURRENT_SEASON}.csv",
                BASE_DIR / "data" / "computed" / f"team_ladders_{CURRENT_SEASON}.csv",
            ]
            existing = [f for f in files_to_check if f.exists()]
            report = diff_report(existing)
            elapsed = time.time() - start
            if report:
                logging.warning(report)
                _notify("AFL Data-Diff Alert", report, is_error=True)
                logging.info(f"  DONE [{name}] with alerts in {elapsed:.1f}s")
            else:
                logging.info(f"  DONE [{name}] no anomalies in {elapsed:.1f}s")
            return True, None  # Advisory only — never blocks pipeline

        else:
            logging.warning(f"  Unknown inline step: {name}")
            return False, f"Unknown inline step: {name}"

    except Exception as e:
        elapsed = time.time() - start
        logging.error(f"  ERROR [{name}] {e}")
        return False, str(e)


# ============================================================================
# STEP RUNNER
# ============================================================================
def run_step(name, script, args, description, python_exe):
    """Run a single update step as a subprocess."""
    # Handle inline steps (no script file)
    if script is None:
        return run_inline_step(name, description)

    script_path = BASE_DIR / script

    if not script_path.exists():
        logging.warning(f"SKIP [{name}] – script not found: {script}")
        return False, "Script not found"

    logging.info(f"START [{name}] {description}")
    logging.info(f"  Running: {python_exe} {script_path} {' '.join(args)}")

    start = time.time()
    try:
        # Some steps need more time
        # - Extended scraper: 18 teams × 25 years of drafts
        # - Traits API: external API call per player (~800 players)
        if "extended" in name or "traits" in name:
            step_timeout = 1200  # 20 minutes
        else:
            step_timeout = 600  # 10 minutes
        result = subprocess.run(
            [python_exe, str(script_path)] + args,
            cwd=str(BASE_DIR),
            capture_output=True,
            text=True,
            timeout=step_timeout,
            env={
                **os.environ,
                "PROTOCOL_BUFFERS_PYTHON_IMPLEMENTATION": "python",
            },
        )
        elapsed = time.time() - start

        if result.returncode == 0:
            logging.info(f"  DONE [{name}] in {elapsed:.1f}s")
            # Log last few lines of stdout for context
            stdout_lines = result.stdout.strip().split("\n")
            for line in stdout_lines[-5:]:
                logging.info(f"    | {line}")
            return True, None
        else:
            logging.error(f"  FAILED [{name}] exit code {result.returncode} ({elapsed:.1f}s)")
            # Log stderr
            if result.stderr:
                for line in result.stderr.strip().split("\n")[-10:]:
                    logging.error(f"    | {line}")
            # Also log stdout tail in case error info is there
            if result.stdout:
                for line in result.stdout.strip().split("\n")[-5:]:
                    logging.error(f"    | {line}")
            return False, f"Exit code {result.returncode}"

    except subprocess.TimeoutExpired:
        elapsed = time.time() - start
        logging.error(f"  TIMEOUT [{name}] after {elapsed:.1f}s")
        return False, f"Timed out (>{step_timeout}s)"

    except Exception as e:
        logging.error(f"  ERROR [{name}] {e}")
        return False, str(e)


# ============================================================================
# MAIN
# ============================================================================
def main():
    parser = argparse.ArgumentParser(
        description="AFL Dashboard scheduled data update"
    )
    parser.add_argument(
        "--quick",
        action="store_true",
        help="Skip slow steps (Wheelo/Selenium, Traits API, photos)",
    )
    parser.add_argument(
        "--dry-run",
        action="store_true",
        help="Show what would run without executing",
    )
    parser.add_argument(
        "--only",
        nargs="+",
        choices=[s[0] for s in UPDATE_STEPS],
        help="Run only specific steps",
    )
    args = parser.parse_args()

    log_file = setup_logging()

    logging.info("=" * 60)
    logging.info("AFL Dashboard – Scheduled Data Update")
    logging.info(f"Date: {datetime.now():%A %d %B %Y %H:%M}")
    logging.info(f"Mode: {'DRY RUN' if args.dry_run else 'QUICK' if args.quick else 'FULL'}")
    logging.info("=" * 60)

    python_exe = get_python()
    logging.info(f"Python: {python_exe}")

    # Filter steps
    steps = UPDATE_STEPS
    if args.only:
        steps = [s for s in steps if s[0] in args.only]
    elif args.quick:
        steps = [s for s in steps if not s[4]]  # Skip slow steps

    logging.info(f"\nSteps to run ({len(steps)}):")
    for name, script, step_args, desc, slow in steps:
        flag = " [SLOW]" if slow else ""
        logging.info(f"  • {name}: {desc}{flag}")
    logging.info("")

    if args.dry_run:
        logging.info("Dry run complete – no scripts executed.")
        return

    # Execute steps
    results = {}
    skipped = {}
    overall_start = time.time()

    for name, script, step_args, desc, slow in steps:
        # ── Dependency check ──────────────────────────────────────
        deps = STEP_DEPENDENCIES.get(name, [])
        failed_deps = [d for d in deps if d in results and not results[d]["success"]]
        if failed_deps:
            reason = f"Skipped: prerequisite(s) failed ({', '.join(failed_deps)})"
            logging.warning(f"  SKIP [{name}] – {reason}")
            results[name] = {"success": False, "error": reason}
            skipped[name] = True
            continue

        success, error = run_step(name, script, step_args, desc, python_exe)
        results[name] = {"success": success, "error": error}

    total_time = time.time() - overall_start

    # Summary
    logging.info("")
    logging.info("=" * 60)
    logging.info("UPDATE SUMMARY")
    logging.info("=" * 60)

    passed = sum(1 for r in results.values() if r["success"])
    failed = sum(1 for n, r in results.items() if not r["success"] and n not in skipped)
    skip_count = len(skipped)

    for name, result in results.items():
        if name in skipped:
            status = f"SKIPPED: {result['error']}"
            icon = "⊘"
        elif result["success"]:
            status = "OK"
            icon = "✓"
        else:
            status = f"FAILED: {result['error']}"
            icon = "✗"
        logging.info(f"  {icon} {name}: {status}")

    logging.info("")
    logging.info(f"Results: {passed} passed, {failed} failed, {skip_count} skipped")
    logging.info(f"Total time: {total_time:.0f}s ({total_time/60:.1f} min)")
    logging.info(f"Log: {log_file}")

    # Log rotation
    rotate_logs()

    if failed > 0:
        msg = f"{failed} step(s) failed. Check log for details."
        logging.warning(f"\n⚠️  {msg}")
        _notify("AFL Dashboard Update", f"⚠️ {msg}", is_error=True)
        sys.exit(1)
    elif skip_count > 0:
        msg = f"Completed with {skip_count} skipped step(s)."
        logging.warning(f"\n⚠️  {msg}")
        _notify("AFL Dashboard Update", msg)
    else:
        logging.info("\n✅ All steps completed successfully!")
        _notify("AFL Dashboard Update", "✅ All steps completed successfully!")


if __name__ == "__main__":
    main()
