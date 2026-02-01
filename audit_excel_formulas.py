"""
Excel Formula Audit Script
Analyzes all Excel files to document formulas and calculations
"""

import pandas as pd
import openpyxl
from openpyxl.utils import get_column_letter
from pathlib import Path
import json

def audit_workbook(filepath, output_file):
    """Audit a workbook for formulas and structure."""
    
    print(f"\n{'='*80}")
    print(f"AUDITING: {filepath}")
    print(f"{'='*80}")
    
    results = {
        "file": str(filepath),
        "sheets": {},
        "formulas": [],
        "summary": {}
    }
    
    # Load with openpyxl to see formulas
    try:
        wb = openpyxl.load_workbook(filepath, data_only=False)
    except Exception as e:
        print(f"Error loading {filepath}: {e}")
        return results
    
    # Also load with pandas to see sheet structure
    xl = pd.ExcelFile(filepath)
    
    print(f"\nSheets found: {len(xl.sheet_names)}")
    for sheet in xl.sheet_names:
        print(f"  - {sheet}")
    
    results["summary"]["sheet_count"] = len(xl.sheet_names)
    results["summary"]["sheet_names"] = xl.sheet_names
    
    # Analyze each sheet
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        sheet_info = {
            "rows": ws.max_row,
            "cols": ws.max_column,
            "formulas": [],
            "formula_count": 0
        }
        
        formula_count = 0
        sample_formulas = []
        
        # Check all cells for formulas (limit to first 30 rows for speed)
        for row in range(1, min(ws.max_row + 1, 31)):
            for col in range(1, min(ws.max_column + 1, 50)):
                cell = ws.cell(row=row, column=col)
                if cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
                    formula_count += 1
                    if len(sample_formulas) < 10:
                        sample_formulas.append({
                            "cell": f"{get_column_letter(col)}{row}",
                            "formula": cell.value
                        })
        
        sheet_info["formula_count"] = formula_count
        sheet_info["sample_formulas"] = sample_formulas
        results["sheets"][sheet_name] = sheet_info
        
        if formula_count > 0:
            print(f"\n  Sheet '{sheet_name}': {formula_count} formulas found")
            for f in sample_formulas[:5]:
                print(f"    {f['cell']}: {f['formula'][:80]}...")
    
    # Check a sample of data rows for structure
    print(f"\n{'='*80}")
    print("DATA STRUCTURE ANALYSIS")
    print(f"{'='*80}")
    
    key_sheets = ['2025 Summary', '2025 Ladders', 'Summary', '2025']
    for sheet_name in key_sheets:
        if sheet_name in xl.sheet_names:
            print(f"\n--- {sheet_name} ---")
            df = xl.parse(sheet_name, header=None, nrows=5)
            print(f"Shape: {df.shape}")
            for i, row in df.iterrows():
                vals = [str(v)[:20] if pd.notna(v) else 'NaN' for v in row.values[:10]]
                print(f"  Row {i}: {vals}")
    
    with open(output_file, 'w') as f:
        json.dump(results, f, indent=2, default=str)
    
    print(f"\nResults saved to: {output_file}")
    return results


def main():
    print("=" * 80)
    print("AFL DASHBOARD - EXCEL FORMULA AUDIT")
    print("=" * 80)
    
    files_to_audit = [
        "AFL Team Ratings.xlsx",
        "AFL Player Ratings.xlsx", 
        "2025 Traits ENRICHED.xlsx",
        "Wheelo_Team_Data.xlsx",
        "Wheelo_Player_Data.xlsx"
    ]
    
    all_results = {}
    
    for filepath in files_to_audit:
        if Path(filepath).exists():
            output = f"audit_{Path(filepath).stem}.json"
            results = audit_workbook(filepath, output)
            all_results[filepath] = results
        else:
            print(f"\nFile not found: {filepath}")
    
    # Generate summary report
    print("\n" + "=" * 80)
    print("SUMMARY REPORT")
    print("=" * 80)
    
    for filepath, results in all_results.items():
        total_formulas = sum(s.get("formula_count", 0) for s in results.get("sheets", {}).values())
        print(f"\n{filepath}:")
        print(f"  Sheets: {results.get('summary', {}).get('sheet_count', 0)}")
        print(f"  Total formulas (first 30 rows): {total_formulas}")
        
        # List sheets with formulas
        sheets_with_formulas = [
            (name, info["formula_count"]) 
            for name, info in results.get("sheets", {}).items() 
            if info.get("formula_count", 0) > 0
        ]
        if sheets_with_formulas:
            print("  Sheets with formulas:")
            for name, count in sorted(sheets_with_formulas, key=lambda x: -x[1])[:10]:
                print(f"    - {name}: {count} formulas")


if __name__ == "__main__":
    main()
